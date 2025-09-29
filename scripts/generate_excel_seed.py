import argparse
import datetime as dt
import json
import uuid
from pathlib import Path

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
EXCEL_DIR = BASE_DIR / 'AES Excel Files'
PURCHASE_ORDER_FILE = next(EXCEL_DIR.glob('DOWREAD-ICS.orders.*.xlsx'))
QUANTITY_SURVEY_FILE = next(EXCEL_DIR.glob('DOWREAD-ICS.QS.*.xlsx'))

PO_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_URL, 'aestrak/purchase-order')
QS_NAMESPACE = uuid.uuid5(uuid.NAMESPACE_URL, 'aestrak/quantity-survey')


def format_date(value):
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str) and value:
        try:
            return dt.datetime.strptime(value, '%Y-%m-%d').date()
        except ValueError:
            return None
    return None


def sql_str(value):
    if value is None:
        return 'NULL'
    if isinstance(value, dt.date):
        return f"'{value.isoformat()}'"
    if isinstance(value, dt.datetime):
        return f"'{value.isoformat()}'"
    if isinstance(value, (int, float)):
        return str(value)
    text = str(value).replace("'", "''")
    return f"'{text}'"


def deterministic_uuid(namespace: uuid.UUID, name: str) -> uuid.UUID:
    return uuid.uuid5(namespace, name)


def load_purchase_orders():
    wb = load_workbook(PURCHASE_ORDER_FILE, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    mapped = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        record = dict(zip(headers, row))
        po_number = str(record.get('Purchase order No.')).strip()
        if not po_number or po_number.lower() == 'none':
            continue
        mapped.append(record)
    return mapped


def load_quantity_surveys():
    wb = load_workbook(QUANTITY_SURVEY_FILE, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    mapped = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None for cell in row):
            continue
        record = dict(zip(headers, row))
        qs_number = record.get('Q.S. number')
        po_number = record.get('Purchase order No.')
        if not qs_number or not po_number:
            continue
        mapped.append(record)
    return mapped


def build_seed_data(org_id: uuid.UUID):
    po_records = load_purchase_orders()
    qs_records = load_quantity_surveys()

    qs_totals = {}
    for record in qs_records:
        po_no = str(record.get('Purchase order No.')).strip()
        total = record.get('TOTAL') or 0
        qs_totals.setdefault(po_no, 0)
        qs_totals[po_no] += float(total or 0)

    purchase_orders = []
    for record in po_records:
        po_no = str(record.get('Purchase order No.')).strip()
        po_id = deterministic_uuid(PO_NAMESPACE, po_no)
        order_value = float(record.get('Order value') or 0)
        total_spent = qs_totals.get(po_no, 0.0)
        remaining_budget = max(order_value - total_spent, 0.0)
        utilization = 0.0
        if order_value:
            utilization = round((total_spent / order_value) * 100, 2)

        purchase_orders.append({
            'id': str(po_id),
            'organization_id': str(org_id),
            'purchase_order_no': po_no,
            'status': record.get('Status') or 'open',
            'company': record.get('Company'),
            'order_short_text': record.get('Order short text'),
            'order_value': order_value,
            'total_spent': round(total_spent, 2),
            'remaining_budget': round(remaining_budget, 2),
            'utilization_percent': utilization,
            'vendor_id': (record.get('Vendor ID') or None),
            'vendor_short_term': (record.get('Short term') or None),
            'work_coordinator_name': record.get('Name') or record.get('Work coordinator'),
            'start_date': format_date(record.get('Start date')),
            'completion_date': format_date(record.get('Date of completion')),
        })

    purchase_order_ids = {po['purchase_order_no']: po['id'] for po in purchase_orders}

    quantity_surveys = []
    for record in qs_records:
        po_no = str(record.get('Purchase order No.')).strip()
        qs_no = str(record.get('Q.S. number')).strip()
        qs_id = deterministic_uuid(QS_NAMESPACE, f'{po_no}:{qs_no}')
        purchase_orders_id = purchase_order_ids.get(po_no)

        contractor_contact = record.get('Contractor contact') or record.get('Contractor Contact')

        quantity_surveys.append({
            'id': str(qs_id),
            'organization_id': str(org_id),
            'purchase_order_id': purchase_orders_id,
            'purchase_order_no': po_no,
            'qs_number': qs_no,
            'quantity_survey_short_text': record.get('Quantity survey short text'),
            'contractor_contact': contractor_contact,
            'vendor_id': record.get('Vendor ID'),
            'total': float(record.get('TOTAL') or 0),
            'created_date': format_date(record.get('CREATED')),
            'transfer_date': format_date(record.get('TRANSFERED')),
            'accepted_date': format_date(record.get('Accepted')),
            'invoice_number': record.get('Invoice number'),
            'invoice_date': format_date(record.get('Invoice Document Date')),
            'accounting_document': record.get('Accounting Document'),
        })

    return purchase_orders, quantity_surveys


def build_insert_sql(table, rows, column_order):
    if not rows:
        return ''
    values_lines = []
    for row in rows:
        values = [sql_str(row.get(column)) for column in column_order]
        values_lines.append(f"({', '.join(values)})")
    columns_sql = ', '.join(column_order)
    values_sql = ',\n  '.join(values_lines)
    return f"INSERT INTO {table} ({columns_sql}) VALUES\n  {values_sql};\n"


def main():
    parser = argparse.ArgumentParser(description='Generate SQL seed data from Excel files.')
    parser.add_argument('--organization-id', required=True, help='Target organization UUID for seeded records.')
    parser.add_argument('--output', default=str(BASE_DIR / 'supabase' / 'excel_seed.sql'), help='Output SQL file path.')
    args = parser.parse_args()

    org_id = uuid.UUID(args.organization_id)

    purchase_orders, quantity_surveys = build_seed_data(org_id)

    po_columns = [
        'id',
        'organization_id',
        'purchase_order_no',
        'status',
        'company',
        'order_short_text',
        'order_value',
        'total_spent',
        'remaining_budget',
        'utilization_percent',
        'vendor_id',
        'vendor_short_term',
        'work_coordinator_name',
        'start_date',
        'completion_date',
    ]

    qs_columns = [
        'id',
        'organization_id',
        'purchase_order_id',
        'purchase_order_no',
        'qs_number',
        'quantity_survey_short_text',
        'contractor_contact',
        'vendor_id',
        'total',
        'created_date',
        'transfer_date',
        'accepted_date',
        'invoice_number',
        'invoice_date',
        'accounting_document',
    ]

    sql_sections = []
    sql_sections.append('-- Auto-generated seed data from Excel files.\n')
    sql_sections.append(f"-- Organization ID: {org_id}\n")
    sql_sections.append(build_insert_sql('public.purchase_orders', purchase_orders, po_columns))
    sql_sections.append(build_insert_sql('public.quantity_surveys', quantity_surveys, qs_columns))

    output_path = Path(args.output)
    output_path.write_text('\n'.join(filter(None, sql_sections)))
    print(f'Wrote seed SQL to {output_path}')


if __name__ == '__main__':
    main()
